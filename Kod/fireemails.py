import smtplib
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import COMMASPACE, formatdate
from email import encoders
import glob
import os 
from shutil import copyfile
from openpyxl import load_workbook

def send_mail(send_to, subject, message, files=[], password=''):
    """Compose and send email with provided info and attachments.
    Args:
        send_to (list[str]): to name(s)
        subject (str): message title
        message (str): message body
        files (list[str]): list of file paths to be attached to email
        password (str): server auth password
    """
    username = 'mortssaerdna2@gmail.com'
    server = 'smtp.gmail.com'
    port = 587
    msg = MIMEMultipart()
    msg['From'] = username
    msg['To'] = COMMASPACE.join(send_to)
    msg['Date'] = formatdate(localtime=True)
    msg['Subject'] = subject

    msg.attach(MIMEText(message))
    for path in files:
        part = MIMEBase('application', "octet-stream")
        with open(path, 'rb') as file:
            part.set_payload(file.read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition',
                        'attachment; filename="{}"'.format(Path(path).name))
        msg.attach(part)

    smtp = smtplib.SMTP(server, port)
    smtp.starttls()
    smtp.login(username, password)
    print("Sending results")
    smtp.sendmail(username, send_to, msg.as_string())
    smtp.quit()
  
def mail_GNNs(send_to, directory, note, password='', send_all=False, seed=None, xlsx=True, send_accs=False):
    dir_path = os.path.dirname(os.path.realpath(__file__))
    path = dir_path+'/'+directory+'results/mailresults.txt'
    path2 = dir_path+'/'+directory+'results/'
    f = open(path, "r")
    res = []
    keys = []
    for x in f:
        res.append(x.split(": ",1)[1].replace('\n',''))
        keys.append(x.split(": ",1)[0])
    dictt = dict(zip(keys, res))
    f.close()
    sub = dictt["model"]+", "+dictt["seed"]+", "+dictt["aggr_func"]+", "+dictt["dataset"]+", "+dictt["date"]
    msg = """
    {}, {}, seed {}, {}, params: {}
    test_acc: {}
    train_acc: {}
    epochs: {}
    avg_time_per_epoch (s): {}
    total_time (h): {}
    note: {}
    """.format(dictt["model"], dictt["aggr_func"], dictt["seed"], dictt["dataset"], dictt["params"], dictt["testacc"], dictt["trainacc"], dictt["epochs"], dictt["avg_time_per_epoch"], dictt["total_time"], note)

    if send_all:
        msg = """
        results appended 
        note: {}
        """.format(note)
        sub = "SUMMARY: "+dictt["model"]+", "+dictt["aggr_func"]+", "+dictt["dataset"]+", "+dictt["date"]
        files = glob.glob(path2 + 'res_*.txt')
    else:
        files = []
        files.append(path)

    if seed is not None:
        copyfile(path, path2 + 'res_' + str(seed) + '.txt')

    if send_accs:
        files.append('accs.mat')
    
    if xlsx:
        import xlsxwriter
        get_next_dict = iter([dictt])
        headers = dictt.keys()
        for key, value in dictt.items():
            dictt[key] = value.replace('.',',')
        if not os.path.isfile(path2+'tmp.csv'):
            with open(path2+'tmp.csv', 'w')as csv_file:
                csv_file.writelines(', '.join(headers))

        if not os.path.isfile(path2+'res_'+dictt["seed"]+'.xlsx'):
            book = xlsxwriter.Workbook(path2+'res_'+dictt["seed"]+'.xlsx')
            sheet = book.add_worksheet("TestSheet")
            for (idx, header) in enumerate(headers):
                sheet.write(0, idx, header)
            book.close()

        with open(path2+'tmp.csv', 'a+') as csv_file:
            book = load_workbook(path2+'res_'+dictt["seed"]+'.xlsx')
            sheet = book.get_sheet_by_name('TestSheet')

            # loop through all dictionaries
            for d in get_next_dict:
                values = [d[key] for key in headers]
                csv_string = '\n'+', '.join(values)
                # write to csv file
                csv_file.write(csv_string)
                # write to excel file
                sheet.append(values)
            book.save(filename=path2+'res_'+dictt["seed"]+'.xlsx')
        files.append(path2+'res_'+dictt["seed"]+'.xlsx')

    send_mail(send_to, sub, msg, files, 'gnns-wagg')